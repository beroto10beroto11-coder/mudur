import io
from typing import Annotated, Any

from fastapi import APIRouter, Depends, HTTPException, Response, Query
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.auth.dependencies import CurrentUser
from app.core.database import get_db
from app.models.timetable import TimetableLesson, Timetable
from app.models.teacher import Teacher
from app.models.course import Course
from app.models.class_ import ClassGroup
from app.models.classroom import Classroom
from app.models.school import School

router = APIRouter(prefix="/exports", tags=["Excel & PDF Exports"])

DAYS = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma"]
PERIODS = [1, 2, 3, 4, 5, 6, 7, 8]


def _try_render_weasyprint_or_html(html_content: str, filename: str) -> Response:
    """Attempt WeasyPrint PDF rendering; fallback to HTML with print script on missing OS libraries."""
    try:
        from weasyprint import HTML
        pdf_bytes = HTML(string=html_content).write_pdf()
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}.pdf"},
        )
    except Exception:
        # Fallback to browser HTML with auto-print script
        printable_html = html_content + "<script>window.onload = function() { window.print(); };</script>"
        return Response(
            content=printable_html,
            media_type="text/html",
        )


def _format_excel_sheet(ws, title: str):
    """Apply styling to an openpyxl worksheet."""
    ws.views.sheetView[0].showGridLines = True

    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    cell_font = Font(name="Calibri", size=10)
    title_font = Font(name="Calibri", size=14, bold=True, color="1E3A8A")

    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    # Apply title
    ws.merge_cells("A1:F1")
    ws["A1"] = title
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    # Header styling (row 3)
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Cell styling
    for row in range(4, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = cell_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Auto-fit column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 15)


async def _load_timetable_data(db: AsyncSession, timetable_id: int):
    t_result = await db.execute(select(Timetable).where(Timetable.id == timetable_id))
    timetable = t_result.scalar_one_or_none()
    if not timetable:
        raise HTTPException(status_code=404, detail="Ders programı bulunamadı.")

    school_res = await db.execute(select(School).where(School.id == timetable.school_id))
    school = school_res.scalar_one_or_none()
    school_name = school.name if school else "Okul Ders Programı"

    lessons_res = await db.execute(select(TimetableLesson).where(TimetableLesson.timetable_id == timetable_id))
    lessons = list(lessons_res.scalars().all())

    teachers = {t.id: t.full_name for t in (await db.execute(select(Teacher))).scalars().all()}
    courses = {c.id: c.name for c in (await db.execute(select(Course))).scalars().all()}
    classes = {c.id: c.name for c in (await db.execute(select(ClassGroup))).scalars().all()}
    rooms = {r.id: r.name for r in (await db.execute(select(Classroom))).scalars().all()}

    return timetable, school_name, lessons, teachers, courses, classes, rooms


# ─────────────────────────────────────────────
# EXCEL EXPORTS
# ─────────────────────────────────────────────

@router.get("/excel/school-schedule")
async def export_school_schedule_excel(
    timetable_id: int,
    db: Annotated[AsyncSession, Depends(get_db)],
    current_user: CurrentUser,
):
    timetable, school_name, lessons, teachers, courses, classes, rooms = await _load_timetable_data(db, timetable_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Okul Programı"

    ws.append([f"{school_name} - {timetable.name}"])
    ws.append([])
    ws.append(["Gün", "Ders Saati", "Sınıf", "Ders", "Öğretmen", "Derslik"])

    for l in lessons:
        ws.append([
            DAYS[l.day] if l.day < len(DAYS) else str(l.day),
            f"{l.period}. Ders",
            classes.get(l.class_id, ""),
            courses.get(l.course_id, ""),
            teachers.get(l.teacher_id, ""),
            rooms.get(l.classroom_id, "-") if l.classroom_id else "-",
        ])

    _format_excel_sheet(ws, f"{school_name} - {timetable.name}")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=Okul_Genel_Programi.xlsx"},
    )


@router.get("/excel/teacher-schedule")
async def export_teacher_schedule_excel(
    timetable_id: int,
    teacher_id: int,
    db: Annotated[AsyncSession, Depends(get_db)],
    current_user: CurrentUser,
):
    timetable, school_name, lessons, teachers, courses, classes, rooms = await _load_timetable_data(db, timetable_id)
    t_name = teachers.get(teacher_id, f"Öğretmen #{teacher_id}")

    t_lessons = [l for l in lessons if l.teacher_id == teacher_id]
    grid = {p: {d: "" for d in range(5)} for p in PERIODS}

    for l in t_lessons:
        c_name = courses.get(l.course_id, "")
        cls_name = classes.get(l.class_id, "")
        grid[l.period][l.day] = f"{c_name}\n({cls_name})"

    wb = Workbook()
    ws = wb.active
    ws.title = t_name[:30]

    ws.append([f"{school_name} - {t_name} Ders Programı"])
    ws.append([])
    ws.append(["Ders Saati"] + DAYS)

    for p in PERIODS:
        row = [f"{p}. Ders"] + [grid[p][d] for d in range(5)]
        ws.append(row)

    _format_excel_sheet(ws, f"{t_name} Ders Programı")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=Ogretmen_{teacher_id}_Program.xlsx"},
    )


@router.get("/excel/class-schedule")
async def export_class_schedule_excel(
    timetable_id: int,
    class_id: int,
    db: Annotated[AsyncSession, Depends(get_db)],
    current_user: CurrentUser,
):
    timetable, school_name, lessons, teachers, courses, classes, rooms = await _load_timetable_data(db, timetable_id)
    cls_name = classes.get(class_id, f"Sınıf #{class_id}")

    c_lessons = [l for l in lessons if l.class_id == class_id]
    grid = {p: {d: "" for d in range(5)} for p in PERIODS}

    for l in c_lessons:
        c_name = courses.get(l.course_id, "")
        t_name = teachers.get(l.teacher_id, "")
        grid[l.period][l.day] = f"{c_name}\n({t_name})"

    wb = Workbook()
    ws = wb.active
    ws.title = cls_name[:30]

    ws.append([f"{school_name} - {cls_name} Ders Programı"])
    ws.append([])
    ws.append(["Ders Saati"] + DAYS)

    for p in PERIODS:
        row = [f"{p}. Ders"] + [grid[p][d] for d in range(5)]
        ws.append(row)

    _format_excel_sheet(ws, f"{cls_name} Ders Programı")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=Sinif_{class_id}_Program.xlsx"},
    )


@router.get("/excel/all-teachers")
async def export_all_teachers_excel(
    timetable_id: int,
    db: Annotated[AsyncSession, Depends(get_db)],
    current_user: CurrentUser,
):
    timetable, school_name, lessons, teachers, courses, classes, rooms = await _load_timetable_data(db, timetable_id)

    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    for t_id, t_name in teachers.items():
        t_lessons = [l for l in lessons if l.teacher_id == t_id]
        if not t_lessons:
            continue

        grid = {p: {d: "" for d in range(5)} for p in PERIODS}
        for l in t_lessons:
            c_name = courses.get(l.course_id, "")
            cls_name = classes.get(l.class_id, "")
            grid[l.period][l.day] = f"{c_name}\n({cls_name})"

        ws = wb.create_sheet(title=t_name[:30].replace("/", "-"))
        ws.append([f"{school_name} - {t_name} Ders Programı"])
        ws.append([])
        ws.append(["Ders Saati"] + DAYS)

        for p in PERIODS:
            ws.append([f"{p}. Ders"] + [grid[p][d] for d in range(5)])

        _format_excel_sheet(ws, f"{t_name} Ders Programı")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Tum_Ogretmenler_Programi.xlsx"},
    )


@router.get("/excel/all-classes")
async def export_all_classes_excel(
    timetable_id: int,
    db: Annotated[AsyncSession, Depends(get_db)],
    current_user: CurrentUser,
):
    timetable, school_name, lessons, teachers, courses, classes, rooms = await _load_timetable_data(db, timetable_id)

    wb = Workbook()
    wb.remove(wb.active)

    for c_id, c_name in classes.items():
        cls_lessons = [l for l in lessons if l.class_id == c_id]
        if not cls_lessons:
            continue

        grid = {p: {d: "" for d in range(5)} for p in PERIODS}
        for l in cls_lessons:
            crs_name = courses.get(l.course_id, "")
            t_name = teachers.get(l.teacher_id, "")
            grid[l.period][l.day] = f"{crs_name}\n({t_name})"

        ws = wb.create_sheet(title=c_name[:30].replace("/", "-"))
        ws.append([f"{school_name} - {c_name} Ders Programı"])
        ws.append([])
        ws.append(["Ders Saati"] + DAYS)

        for p in PERIODS:
            ws.append([f"{p}. Ders"] + [grid[p][d] for d in range(5)])

        _format_excel_sheet(ws, f"{c_name} Ders Programı")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Tum_Siniflar_Programi.xlsx"},
    )


# ─────────────────────────────────────────────
# PDF / A4 PRINTABLE HTML EXPORTS
# ─────────────────────────────────────────────

def _generate_schedule_html(title: str, subtitle: str, tables_data: list[dict[str, Any]]) -> str:
    """Generate clean, responsive A4 Landscape HTML with print CSS styling."""
    html_sections = []

    for tdata in tables_data:
        headers_html = "".join(f"<th>{h}</th>" for h in tdata["headers"])
        rows_html = ""
        for row in tdata["rows"]:
            cells_html = "".join(f"<td>{cell}</td>" for cell in row)
            rows_html += f"<tr>{cells_html}</tr>"

        html_sections.append(f"""
        <div class="schedule-card">
            <h2>{tdata['card_title']}</h2>
            <table>
                <thead><tr>{headers_html}</tr></thead>
                <tbody>{rows_html}</tbody>
            </table>
        </div>
        """)

    full_html = f"""<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<title>{title}</title>
<style>
    @page {{
        size: A4 landscape;
        margin: 10mm;
    }}
    body {{
        font-family: 'Segoe UI', Roboto, Helvetica, Arial, sans-serif;
        color: #1e293b;
        margin: 0;
        padding: 0;
        background: #ffffff;
    }}
    .header {{
        text-align: center;
        margin-bottom: 15px;
        border-bottom: 2px solid #1e3a8a;
        padding-bottom: 8px;
    }}
    .header h1 {{
        margin: 0;
        font-size: 20px;
        color: #1e3a8a;
        text-transform: uppercase;
    }}
    .header p {{
        margin: 4px 0 0 0;
        font-size: 12px;
        color: #64748b;
    }}
    .grid-container {{
        display: flex;
        flex-direction: column;
        gap: 20px;
    }}
    .schedule-card {{
        page-break-inside: avoid;
        margin-bottom: 20px;
    }}
    .schedule-card h2 {{
        font-size: 14px;
        margin: 0 0 8px 0;
        color: #0f172a;
        border-left: 4px solid #2563eb;
        padding-left: 8px;
    }}
    table {{
        width: 100%;
        border-collapse: collapse;
        font-size: 11px;
        text-align: center;
    }}
    th {{
        background-color: #1e3a8a;
        color: #ffffff;
        font-weight: 600;
        padding: 6px;
        border: 1px solid #0f172a;
        text-transform: uppercase;
    }}
    td {{
        border: 1px solid #cbd5e1;
        padding: 6px;
        height: 38px;
        vertical-align: middle;
        white-space: pre-line;
    }}
    tr:nth-child(even) td {{
        background-color: #f8fafc;
    }}
    .footer {{
        margin-top: 30px;
        display: flex;
        justify-content: space-between;
        font-size: 11px;
        color: #64748b;
    }}
</style>
</head>
<body>
    <div class="header">
        <h1>{title}</h1>
        <p>{subtitle}</p>
    </div>
    <div class="grid-container">
        {"".join(html_sections)}
    </div>
    <div class="footer">
        <span>Tarih: {pd.Timestamp.now().strftime('%d.%m.%Y')}</span>
        <span>Okul Yönetim & Otomatik Dağıtım Sistemi</span>
        <span>İmza / Mühür</span>
    </div>
</body>
</html>
"""
    return full_html


@router.get("/pdf/school-schedule")
async def export_school_schedule_pdf(
    timetable_id: int,
    db: Annotated[AsyncSession, Depends(get_db)],
    current_user: CurrentUser,
):
    timetable, school_name, lessons, teachers, courses, classes, rooms = await _load_timetable_data(db, timetable_id)

    headers = ["Ders Saati"] + DAYS
    rows = []

    grid = {p: {d: "" for d in range(5)} for p in PERIODS}
    for l in lessons:
        c_name = courses.get(l.course_id, "")
        cls_name = classes.get(l.class_id, "")
        t_name = teachers.get(l.teacher_id, "")
        grid[l.period][l.day] += f"{cls_name}: {c_name} ({t_name})\n"

    for p in PERIODS:
        rows.append([f"{p}. Ders"] + [grid[p][d].strip() for d in range(5)])

    html = _generate_schedule_html(
        title=school_name,
        subtitle=f"{timetable.name} - Genel Okul Ders Programı",
        tables_data=[{
            "card_title": "Tüm Okul Ders Programı Matrisi",
            "headers": headers,
            "rows": rows,
        }],
    )
    return _try_render_weasyprint_or_html(html, "Okul_Genel_Programi")


@router.get("/pdf/teacher-schedule")
async def export_teacher_schedule_pdf(
    timetable_id: int,
    db: Annotated[AsyncSession, Depends(get_db)],
    current_user: CurrentUser,
    teacher_id: int | None = None,
):
    timetable, school_name, lessons, teachers, courses, classes, rooms = await _load_timetable_data(db, timetable_id)

    target_teachers = [teacher_id] if teacher_id else list(teachers.keys())
    tables_data = []

    for t_id in target_teachers:
        t_name = teachers.get(t_id, f"Öğretmen #{t_id}")
        t_lessons = [l for l in lessons if l.teacher_id == t_id]
        if not t_lessons and not teacher_id:
            continue

        grid = {p: {d: "" for d in range(5)} for p in PERIODS}
        for l in t_lessons:
            c_name = courses.get(l.course_id, "")
            cls_name = classes.get(l.class_id, "")
            grid[l.period][l.day] = f"{c_name}\n({cls_name})"

        rows = []
        for p in PERIODS:
            rows.append([f"{p}. Ders"] + [grid[p][d] for d in range(5)])

        tables_data.append({
            "card_title": f"Öğretmen: {t_name}",
            "headers": ["Ders Saati"] + DAYS,
            "rows": rows,
        })

    html = _generate_schedule_html(
        title=school_name,
        subtitle=f"{timetable.name} - Öğretmen Ders Programları",
        tables_data=tables_data,
    )
    return _try_render_weasyprint_or_html(html, "Ogretmen_Ders_Programlari")


@router.get("/pdf/class-schedule")
async def export_class_schedule_pdf(
    timetable_id: int,
    db: Annotated[AsyncSession, Depends(get_db)],
    current_user: CurrentUser,
    class_id: int | None = None,
):
    timetable, school_name, lessons, teachers, courses, classes, rooms = await _load_timetable_data(db, timetable_id)

    target_classes = [class_id] if class_id else list(classes.keys())
    tables_data = []

    for c_id in target_classes:
        cls_name = classes.get(c_id, f"Sınıf #{c_id}")
        cls_lessons = [l for l in lessons if l.class_id == c_id]
        if not cls_lessons and not class_id:
            continue

        grid = {p: {d: "" for d in range(5)} for p in PERIODS}
        for l in cls_lessons:
            c_name = courses.get(l.course_id, "")
            t_name = teachers.get(l.teacher_id, "")
            grid[l.period][l.day] = f"{c_name}\n({t_name})"

        rows = []
        for p in PERIODS:
            rows.append([f"{p}. Ders"] + [grid[p][d] for d in range(5)])

        tables_data.append({
            "card_title": f"Sınıf: {cls_name}",
            "headers": ["Ders Saati"] + DAYS,
            "rows": rows,
        })

    html = _generate_schedule_html(
        title=school_name,
        subtitle=f"{timetable.name} - Sınıf Ders Programları",
        tables_data=tables_data,
    )
    return _try_render_weasyprint_or_html(html, "Sinif_Ders_Programlari")
